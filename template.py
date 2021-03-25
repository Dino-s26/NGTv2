import os
import paramiko
import time
import datetime
import socket
from openpyxl import load_workbook
from pathlib import Path

''' Hard code source file *Recommended method '''
filename = "x://path_to_file.xlsx" ''' *change this to your excel file '''
''' Should you choose to do input filename manually '''
#filename = input("Full path to excel file : ")

def get_date():
    #Define date & time
    '''
    ddt for get date time with format, example of the format : 18-11-20--20-21-21
    c_month for get month and year format, example of the format : November-2020
    '''
    ddt = str(datetime.datetime.now().strftime("%d-%m-%y--%H-%M-%M"))
    c_month = str(datetime.datetime.now().strftime('%B-%Y'))
    return ddt, c_month

def create_log():
    '''
    Create log for script for debug purpose
    '''
    path = "C:/log/"
    c_log = Path(path).mkdir(parents=True, exist_ok=True)
    log_name = "script_log-[Device]_Capture-"+get_date()[0]+".log" ''' change [Device] to appropriate device use '''
    return path, c_log, log_name

def output_log():
    '''
    Create output log folder for result of the script 
    '''
    path_output = "C:/output_log/"
    o_output = Path(path_output).mkdir(parents=True, exist_ok=True)
    return path_output, o_output

def g_read_credential():
    wb = load_workbook(filename, read_only=True)
    creds = wb['credentials']
    for value in creds.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        '''
        for debug purpose uncomment below command
        '''
        #print("Username : ",value[0],"\nPassword : ", value[1])
        '''Return Value of the credential '''
        return value[0], value[1]

def ssh_method():
    wb = load_workbook(filename, read_only=True)
    ip = wb['list_ip']
    max_rw = ip.max_row
    max_cl = ip.max_column
    ''' for debug number of ROW and COLUMN in excel '''
    #print("ROW: ",max_rw, "COLUMN: ",max_cl)
    for value_ip in range(2, max_rw+1):
        '''
        for debug purpose, change the value of the column
        '''
        lip = ip.cell(row=value_ip, column=2).value
        lhost = ip.cell(row=value_ip, column=1).value
        ''' for debug list of ip to SSH in excel sheet of list_ip '''
        #p_ip = "ssh "+lip+"\n"
        #print(p_ip)

        ''' Old iteration method, just for debug purpose incase the newer iteration not working '''
    #for host in ip :
        '''
        SSH activity new method to pass exception
        so that the fault can be skipped and logged to a .log file defined on create_log() function
        '''
        try:
            pre_conn = paramiko.SSHClient()
            pre_conn.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            pre_conn.connect(lip, port=22, username=g_read_credential()[0], password=g_read_credential()[1], allow_agent=False, timeout=10)
            print("Establishing SSH Connectiont to : ", lip)
            conn = pre_conn.invoke_shell()
            '''
            Change the format below for the log_name file
            '''
            log_name = lhost + '--' + lip +"--"+ get_date()[0] + '.txt'
            done = "SSH Done for Hostname : "+lhost+", with IP : "+lip
            print(done)
            c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
            c_logg.write("--[Success Log]--\n")
            c_logg.write("["+get_date()[0]+"] - ")
            c_logg.write(done+"\n")
            c_logg.close()

            wb = load_workbook(filename, read_only=True)
            cmd = wb['commands']
            max_rw = cmd.max_row
            max_cl = cmd.max_column

            for command in cmd.iter_rows(min_row=2, max_row=max_rw, min_col=1, max_col=max_cl, values_only=True):
                ''' debug for command defined on excel file for commands sheet '''
                #print("command are : ",''.join(command),"\n")
                
                cmds = ''.join(command)+"\n"
                ''' debug purpose for command sheet '''
                #print(cmds)

                conn.send(cmds)
                time.sleep(5)
                recv_buff = conn.recv(65000)
                
                ''' debug if the command are executed '''
                #print(''.join(command), "Done !")
                
                output_file = open(os.path.join(output_log()[0],log_name), 'ab+')
                output_file.write(recv_buff)
                output_file.close
            conn.close()
            print ("All commands are capture and executed for :", lip,"\n")

        '''
        except timeout socket, skipped it and logged it to the log file defined
        '''
        except socket.timeout as timout:
            exception = "SSH skip for the Host : "+lip+", Host unreachable / not available.\n"
            print(exception)
            c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
            c_logg.write("--[Skip Host Log]--\n")
            c_logg.write("["+get_date()[0]+"] - ")
            c_logg.write(exception+"\n")
            c_logg.close()
            pass

        '''
        except invalid ip format, skipped it and logged it to the log file defined
        '''
        except socket.gaierror as invalid_ip:
            exception = "Invalid IP Format for : "+lip+", Check your IP address again.\n"
            print(exception)
            c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
            c_logg.write("--[Invalid IP Format Log]--\n")
            c_logg.write("["+get_date()[0]+"] - ")
            c_logg.write(exception+"\n")
            c_logg.close()
            pass
        
        '''
        except invalid ip format by unicode, skipped it and logged it to the log file defined
        '''
        except UnicodeError as invalid_ip_format:
            exception = "Invalid IP Format for : "+lip+", Check your IP address again.\n"
            print(exception)
            c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
            c_logg.write("--[Invalid IP Format Log]--\n")
            c_logg.write("["+get_date()[0]+"] - ")
            c_logg.write(exception+"\n")
            c_logg.close()
            pass

        '''
        except authentication problem, skipped it and logged it to the log file defined
        '''
        except paramiko.ssh_exception.AuthenticationException as invalid_auth:
            exception = "Invalid Authentication for : "+lipt+", Check Username and/or Password !\n"
            print(exception)
            c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
            c_logg.write("--[Invalid Authentication Log]--\n")
            c_logg.write("["+get_date()[0]+"] - ")
            c_logg.write(exception+"\n")
            c_logg.close()
            pass

        '''
        The exception below is used to skip any TypeError traceback,
        \n this should not be used in production, just for troubleshooting purpose.
        \n so leave it on comment 
        '''

        #except TypeError as t_error:
        #    exception = "Type Error Exception detected ! \nkindly run the script manualy to check the traceback, if this a mistake you can ignore it.\n"
        #    print (exception)
        #    c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
        #    c_logg.write("--[Type Error Log]--\n")
        #    c_logg.write("["+get_date()[0]+"] - ")
        #    c_logg.write(exception+" "+lip+"\n")
        #    c_logg.close()
        #    pass

        '''
        The exception below is used when excel file detect any empty value that previously
        have value, should you delete value in list_ip sheet, make sure use clear all for the
        selected column that you deleted
        '''

        except paramiko.ssh_exception.NoValidConnectionsError as invalid_connection:
            exception = "Invalid connection detected !, ignoring invalid connection.\n"
            print (exception)
            c_logg = open(os.path.join(create_log()[0],create_log()[2]), 'a')
            c_logg.write("--[Invalid Connection Log]--\n")
            c_logg.write("["+get_date()[0]+"] - ")
            c_logg.write(exception+"\n")
            c_logg.close()
            pass
'''
Just to show it on terminal / cmd, if you don't need it you just need to remove it from print() function.
'''
print(ssh_method())
